from openpyxl import load_workbook


def replace_xls(excel_name1, excel_name2, sheet_name):
    '''
    用python实现跨excel工作表sheet之间的复制
    excel_name1为旧表，excel_name2为新表，需要自行刷格式,sheetname为待复制的sheet
    '''
    wb1 = load_workbook(excel_name1)
    wb2 = load_workbook(excel_name2)
    for k in sheet_name:
        ws1 = wb1[k]
        ws2 = wb2[k]
        for i, row in enumerate(ws1.iter_rows()):
            for j, cell in enumerate(row):
                ws2.cell(row=i+1, column=j+1, value=cell.value)
    wb2.save(excel_name2)


print('debug0')
excel_name = r'C:\Users\hljge\git_folder\test1.xlsx'
print('debug1')
sheet_names = [u'Sheet2', u'Sheet1']
# print('debug2')
# test_excel = 'E:\\04-债券\\13-债基\\业绩分析\\基础数据20200528.xlsx'
print('debug3')
target_excel = 'C:\\Users\\hljge\\git_folder\\test2.xlsx'
print('debug4')
replace_xls(excel_name, target_excel, sheet_names)

